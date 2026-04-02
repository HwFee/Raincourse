"""
题目导出服务
支持导出题目到Excel、CSV、JSON等格式
"""
import json
import os
import csv
from datetime import datetime
from typing import List, Dict, Optional
from utils.utils import get_project_root


class QuestionExporter:
    """题目导出器"""
    
    def __init__(self):
        self.base_dir = get_project_root()
        self.export_dir = os.path.join(self.base_dir, "exports")
        os.makedirs(self.export_dir, exist_ok=True)
    
    def export_to_json(self, questions: List[Dict], filename: str = None) -> Dict:
        """导出为JSON格式"""
        try:
            if not filename:
                filename = f"questions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            
            if not filename.endswith('.json'):
                filename += '.json'
            
            filepath = os.path.join(self.export_dir, filename)
            
            export_data = {
                "export_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "total_count": len(questions),
                "questions": questions
            }
            
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, ensure_ascii=False, indent=2)
            
            return {
                "success": True,
                "message": "导出成功",
                "filepath": filepath,
                "count": len(questions)
            }
        except Exception as e:
            return {
                "success": False,
                "message": f"导出失败: {str(e)}"
            }
    
    def export_to_csv(self, questions: List[Dict], filename: str = None) -> Dict:
        """导出为CSV格式"""
        try:
            if not filename:
                filename = f"questions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            
            if not filename.endswith('.csv'):
                filename += '.csv'
            
            filepath = os.path.join(self.export_dir, filename)
            
            # 准备CSV数据
            csv_data = []
            for idx, question in enumerate(questions, 1):
                row = {
                    "序号": idx,
                    "题目ID": question.get("ProblemID", ""),
                    "题型": question.get("TypeText", ""),
                    "分值": question.get("Score", 0),
                    "题目内容": self._clean_html(question.get("Body", "")),
                    "选项": self._format_options(question.get("Options", [])),
                    "答案": question.get("Answer", ""),
                    "解析": question.get("Analysis", "")
                }
                csv_data.append(row)
            
            # 写入CSV
            if csv_data:
                with open(filepath, 'w', encoding='utf-8-sig', newline='') as f:
                    writer = csv.DictWriter(f, fieldnames=csv_data[0].keys())
                    writer.writeheader()
                    writer.writerows(csv_data)
            
            return {
                "success": True,
                "message": "导出成功",
                "filepath": filepath,
                "count": len(questions)
            }
        except Exception as e:
            return {
                "success": False,
                "message": f"导出失败: {str(e)}"
            }
    
    def export_to_excel(self, questions: List[Dict], filename: str = None) -> Dict:
        """导出为Excel格式"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            if not filename:
                filename = f"questions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'
            
            filepath = os.path.join(self.export_dir, filename)
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "题目列表"
            
            # 设置表头
            headers = ["序号", "题目ID", "题型", "分值", "题目内容", "选项", "答案", "解析"]
            ws.append(headers)
            
            # 设置表头样式
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 添加数据
            for idx, question in enumerate(questions, 1):
                row = [
                    idx,
                    question.get("ProblemID", ""),
                    question.get("TypeText", ""),
                    question.get("Score", 0),
                    self._clean_html(question.get("Body", "")),
                    self._format_options(question.get("Options", [])),
                    question.get("Answer", ""),
                    question.get("Analysis", "")
                ]
                ws.append(row)
            
            # 调整列宽
            column_widths = [8, 15, 12, 8, 50, 40, 20, 40]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
            
            # 保存文件
            wb.save(filepath)
            
            return {
                "success": True,
                "message": "导出成功",
                "filepath": filepath,
                "count": len(questions)
            }
        except ImportError:
            return {
                "success": False,
                "message": "未安装openpyxl库,无法导出Excel格式。请运行: pip install openpyxl"
            }
        except Exception as e:
            return {
                "success": False,
                "message": f"导出失败: {str(e)}"
            }
    
    def export_to_markdown(self, questions: List[Dict], filename: str = None) -> Dict:
        """导出为Markdown格式"""
        try:
            if not filename:
                filename = f"questions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.md"
            
            if not filename.endswith('.md'):
                filename += '.md'
            
            filepath = os.path.join(self.export_dir, filename)
            
            lines = []
            lines.append(f"# 题目导出")
            lines.append(f"\n导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            lines.append(f"题目总数: {len(questions)}")
            lines.append("\n---\n")
            
            for idx, question in enumerate(questions, 1):
                lines.append(f"## {idx}. {question.get('TypeText', '未知题型')} ({question.get('Score', 0)}分)")
                lines.append(f"\n**题目ID**: {question.get('ProblemID', '')}")
                lines.append(f"\n**题目内容**:\n\n{self._clean_html(question.get('Body', ''))}")
                
                options = question.get("Options", [])
                if options:
                    lines.append("\n**选项**:\n")
                    for opt in options:
                        lines.append(f"- {opt}")
                
                answer = question.get("Answer", "")
                if answer:
                    lines.append(f"\n**答案**: {answer}")
                
                analysis = question.get("Analysis", "")
                if analysis:
                    lines.append(f"\n**解析**: {analysis}")
                
                lines.append("\n---\n")
            
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write('\n'.join(lines))
            
            return {
                "success": True,
                "message": "导出成功",
                "filepath": filepath,
                "count": len(questions)
            }
        except Exception as e:
            return {
                "success": False,
                "message": f"导出失败: {str(e)}"
            }
    
    def _clean_html(self, text: str) -> str:
        """清理HTML标签"""
        import re
        if not text:
            return ""
        # 移除HTML标签
        clean = re.sub(r'<[^>]+>', '', text)
        # 替换HTML实体
        clean = clean.replace('&nbsp;', ' ')
        clean = clean.replace('&lt;', '<')
        clean = clean.replace('&gt;', '>')
        clean = clean.replace('&amp;', '&')
        # 移除多余空白
        clean = re.sub(r'\s+', ' ', clean).strip()
        return clean
    
    def _format_options(self, options: List) -> str:
        """格式化选项"""
        if not options:
            return ""
        if isinstance(options, list):
            return " | ".join(str(opt) for opt in options)
        return str(options)
    
    def get_export_history(self) -> List[Dict]:
        """获取导出历史"""
        try:
            files = []
            for filename in os.listdir(self.export_dir):
                filepath = os.path.join(self.export_dir, filename)
                if os.path.isfile(filepath):
                    stat = os.stat(filepath)
                    files.append({
                        "filename": filename,
                        "filepath": filepath,
                        "size": f"{stat.st_size / 1024:.2f} KB",
                        "date": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
                    })
            
            # 按时间倒序排列
            files.sort(key=lambda x: x['date'], reverse=True)
            return files
        except Exception as e:
            return []
